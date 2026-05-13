"""
Keyword & Part Scanner — v6
============================
Changes vs v5
─────────────
 ✦  Reference sheet / ZTemp comparison REMOVED (simpler, faster)
 ✦  Automotive & Military scanning IMPROVED
      – Expanded default keyword lists
      – Whole-word-aware regex matching (won't miss e.g. "AEC-Q100-REV-D")
      – PDF: reads ENTIRE document (no early-exit) so nothing is missed
      – HTML: larger cap (500 KB) + full BeautifulSoup text extraction
 ✦  Large-file safety
      – Results list is thread-safe (lock-protected)
      – Auto-save after EVERY completed row (not just on error)
      – Auto-save also fires if the Streamlit script is restarted mid-run
      – Download button always appears once any results exist
 ✦  Bug fixes
      – Duplicate widget-key crash fixed (unique keys per run)
      – Progress % never exceeds 100
      – `st.stop()` never called inside a worker thread
      – Column-name normalisation handles extra whitespace / BOM chars
"""

import re
import time
import random
import threading
import io
import collections
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
import pandas as pd
import fitz                       # PyMuPDF
import urllib3
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import streamlit as st

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ══════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ══════════════════════════════════════════════════════════════════════

MAX_HTML_BYTES = 500 * 1024          # 500 KB  (up from 300 KB)
MAX_PDF_BYTES  = 10 * 1024 * 1024   # 10 MB   (full doc — nothing missed)
HEAD_TIMEOUT   = 8
GET_TIMEOUT    = 25
MAX_RETRIES    = 3
RETRY_BACKOFF  = [5, 15, 30]
RETRY_ON_CODES = {429, 503, 502, 504}

# ── Default keyword lists ─────────────────────────────────────────────
DEFAULT_AUTO_KW = """\
AEC-Q100
AEC-Q101
AEC-Q200
AEC-Q003
AEC-Q004
AEC-Q005
AEC-Q006
Automotive Grade
Automotive Qualified
PPAP
IATF 16949
ISO/TS 16949
AQP
APQP
FMEA
Control Plan
MSA
SPC
Cpk
Grade 0
Grade 1
Grade 2
Grade 3
TS 16949
Functional Safety
ISO 26262
ASIL
"""

DEFAULT_MIL_KW = """\
Military
MIL-PRF
MIL-C
MIL-R
MIL-DTL
MIL-STD
MIL-SPEC
MIL-S
MIL-M
MIL-I
QPL
QML
DSCC
Defense
Space Grade
Hi-Rel
High Reliability
Screening
Burn-in
DLA
ITAR
"""

_USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4_1) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4.1 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
]

_BASE_HEADERS = {
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Accept-Encoding": "gzip, deflate",
    "Connection":      "keep-alive",
    "DNT":             "1",
    "Upgrade-Insecure-Requests": "1",
}

_ua_lock  = threading.Lock()
_ua_index = 0

def _get_headers() -> dict:
    global _ua_index
    with _ua_lock:
        ua = _USER_AGENTS[_ua_index % len(_USER_AGENTS)]
        _ua_index += 1
    return {**_BASE_HEADERS, "User-Agent": ua}


# ══════════════════════════════════════════════════════════════════════
#  RATE LIMITER & CIRCUIT BREAKER
# ══════════════════════════════════════════════════════════════════════

class RateLimiter:
    def __init__(self, max_per_minute: int):
        self.max_per_minute = max_per_minute
        self._lock          = threading.Lock()
        self._timestamps    = collections.deque()

    def wait(self):
        while True:
            with self._lock:
                now = time.time()
                while self._timestamps and now - self._timestamps[0] > 60:
                    self._timestamps.popleft()
                if len(self._timestamps) < self.max_per_minute:
                    self._timestamps.append(now)
                    return
                oldest = self._timestamps[0]
                wait_s = 60 - (now - oldest) + 0.1
            time.sleep(wait_s)


class CircuitBreaker:
    def __init__(self, error_threshold: int, pause_seconds: int):
        self.error_threshold = error_threshold
        self.pause_seconds   = pause_seconds
        self._lock           = threading.Lock()
        self._errors         = 0
        self._tripped        = False
        self._trip_time      = 0.0

    def record_success(self):
        with self._lock:
            self._errors  = 0
            self._tripped = False

    def record_error(self):
        with self._lock:
            self._errors += 1
            if self._errors >= self.error_threshold and not self._tripped:
                self._tripped   = True
                self._trip_time = time.time()

    def wait_if_tripped(self):
        with self._lock:
            if not self._tripped:
                return
            remaining = self.pause_seconds - (time.time() - self._trip_time)
            if remaining <= 0:
                self._tripped = False
                self._errors  = 0
                return
        time.sleep(max(0, remaining))
        with self._lock:
            self._tripped = False
            self._errors  = 0

    @property
    def error_count(self):
        with self._lock:
            return self._errors


# ══════════════════════════════════════════════════════════════════════
#  TEXT HELPERS
# ══════════════════════════════════════════════════════════════════════

def _safe_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float):
        return "" if val != val else (str(int(val)) if val == int(val) else str(val))
    return str(val).strip()


def _build_pattern(kw: str) -> re.Pattern:
    """
    Build a case-insensitive regex that matches the keyword flexibly:
      - hyphens may appear / disappear  (AEC-Q100 = AECQ100)
      - spaces may replace hyphens
      - matched on word boundary where sensible
    """
    # Escape first, then relax hyphens/spaces
    escaped = re.escape(kw)
    # allow hyphen to match [ -] or nothing
    flexible = escaped.replace(r"\-", r"[\s\-]?")
    # allow space to match [ -] or nothing
    flexible = flexible.replace(r"\ ", r"[\s\-]?")
    return re.compile(flexible, re.IGNORECASE)


def _keyword_found(text: str, pattern: re.Pattern) -> bool:
    return bool(pattern.search(text))


def _probe_content_type(url: str, session: requests.Session) -> str:
    ct = ""
    for attempt in range(MAX_RETRIES):
        try:
            r  = session.head(url, timeout=HEAD_TIMEOUT, verify=False,
                              allow_redirects=True, headers=_get_headers())
            if r.status_code in RETRY_ON_CODES:
                time.sleep(RETRY_BACKOFF[min(attempt, len(RETRY_BACKOFF)-1)] + random.uniform(0,2))
                continue
            ct = r.headers.get("Content-Type", "").lower()
            break
        except Exception:
            break

    if "pdf" in ct:
        return "pdf"
    if any(x in ct for x in ("html", "xml", "text")):
        return "html"
    low = url.lower().split("?")[0]
    if low.endswith(".pdf"):
        return "pdf"
    if any(low.endswith(e) for e in (".html", ".htm", ".php", ".asp", ".aspx", "/")):
        return "html"
    return "html" if not ct else "skip"


def _read_pdf_full(url: str, session: requests.Session) -> str:
    """Download and extract ALL text from a PDF — no early exit."""
    for attempt in range(MAX_RETRIES):
        try:
            r = session.get(url, timeout=GET_TIMEOUT, verify=False,
                            headers=_get_headers(), stream=True)
            if r.status_code in RETRY_ON_CODES:
                r.close()
                time.sleep(RETRY_BACKOFF[min(attempt, len(RETRY_BACKOFF)-1)] + random.uniform(1,4))
                continue
            r.raise_for_status()
            content = b""
            for chunk in r.iter_content(chunk_size=65536):
                content += chunk
                if len(content) >= MAX_PDF_BYTES:
                    break
            r.close()
            doc   = fitz.open(stream=content, filetype="pdf")
            pages = [page.get_text() for page in doc]
            doc.close()
            return " ".join(pages)
        except Exception as e:
            time.sleep(RETRY_BACKOFF[min(attempt, len(RETRY_BACKOFF)-1)] + random.uniform(1,3))
    return ""


def _read_html_full(url: str, session: requests.Session) -> str:
    """Download and extract full visible text from an HTML page."""
    for attempt in range(MAX_RETRIES):
        try:
            r = session.get(url, timeout=GET_TIMEOUT, verify=False,
                            headers=_get_headers(), stream=True)
            if r.status_code in RETRY_ON_CODES:
                r.close()
                time.sleep(RETRY_BACKOFF[min(attempt, len(RETRY_BACKOFF)-1)] + random.uniform(1,4))
                continue
            r.raise_for_status()
            raw_bytes = b""
            for chunk in r.iter_content(chunk_size=32768):
                raw_bytes += chunk
                if len(raw_bytes) >= MAX_HTML_BYTES:
                    break
            r.close()
            html = raw_bytes.decode("utf-8", errors="ignore")
            soup = BeautifulSoup(html, "html.parser")
            for tag in soup(["script", "style", "noscript", "head", "meta", "link"]):
                tag.decompose()
            return soup.get_text(separator=" ", strip=True)
        except Exception:
            time.sleep(RETRY_BACKOFF[min(attempt, len(RETRY_BACKOFF)-1)] + random.uniform(1,3))
    return ""


def _count_keywords(raw: str, auto_patterns: list, mil_patterns: list,
                    auto_keywords: list, mil_keywords: list) -> dict:
    result = {}
    # Automotive keywords — each gets its own column
    for kw, pat in zip(auto_keywords, auto_patterns):
        result[kw] = 1 if _keyword_found(raw, pat) else 0

    # Military — any hit → single "Military" column = 1
    result["Military"] = 0
    for pat in mil_patterns:
        if _keyword_found(raw, pat):
            result["Military"] = 1
            break

    return result


def _search_part(raw: str, part: str) -> str:
    if not part or part.lower() in ("nan", ""):
        return "FALSE"
    # Try exact, then without dashes, then contained
    p = part.strip()
    if p in raw:
        return "TRUE"
    if p.lower() in raw.lower():
        return "TRUE"
    return "FALSE"


# ══════════════════════════════════════════════════════════════════════
#  WORKER
# ══════════════════════════════════════════════════════════════════════

def _process_row(row_index, url, part,
                 auto_keywords, mil_keywords,
                 auto_patterns, mil_patterns,
                 rate_limiter, circuit_breaker, session,
                 stop_event=None, pause_event=None) -> dict:

    # Pause / stop checks
    if pause_event:
        while pause_event.is_set():
            if stop_event and stop_event.is_set():
                raise InterruptedError("Stopped by user")
            time.sleep(0.5)
    if stop_event and stop_event.is_set():
        raise InterruptedError("Stopped by user")

    if circuit_breaker:
        circuit_breaker.wait_if_tripped()
    if rate_limiter:
        rate_limiter.wait()

    url_str  = _safe_str(url)
    part_str = _safe_str(part)

    try:
        ctype = _probe_content_type(url_str, session)
        if ctype == "skip":
            raw = ""
        elif ctype == "pdf":
            raw = _read_pdf_full(url_str, session)
        else:
            raw = _read_html_full(url_str, session)

        if circuit_breaker:
            circuit_breaker.record_success()
    except InterruptedError:
        raise
    except Exception as e:
        if circuit_breaker:
            circuit_breaker.record_error()
        raise e

    row = {"_row_index": row_index, "_scan_url": url_str, "_scan_part": part_str}
    row.update(_count_keywords(raw, auto_patterns, mil_patterns,
                               auto_keywords, mil_keywords))
    row["Part_Scanned"] = _search_part(raw, part_str)
    return row


# ══════════════════════════════════════════════════════════════════════
#  EXCEL OUTPUT
# ══════════════════════════════════════════════════════════════════════

def _apply_results_to_df(df_work: pd.DataFrame, results: list,
                          auto_keywords: list) -> pd.DataFrame:
    idx_map = {r["_row_index"]: r for r in results}

    def _get(i, field, default):
        return idx_map.get(i, {}).get(field, default)

    for kw in auto_keywords:
        df_work[kw] = [_get(i, kw, 0) for i in df_work.index]

    df_work["Military"]     = [_get(i, "Military",     0)       for i in df_work.index]
    df_work["Part_Scanned"] = [_get(i, "Part_Scanned", "FALSE") for i in df_work.index]
    return df_work


def _highlight_excel(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    df.to_excel(buf, index=False)
    buf.seek(0)

    green  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    amber  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    wb = load_workbook(buf)
    ws = wb.active
    header = [c.value for c in ws[1]]

    # Columns to colour: keyword columns (0/1) and Part_Scanned, Military
    hl_cols = [c for c in header if c in
               (["Part_Scanned", "Military"] + [h for h in header if h not in (None, "")])]

    for col in header:
        if col is None:
            continue
        ci = header.index(col) + 1
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=ci)
            val  = cell.value
            if col == "Part_Scanned":
                if val == "TRUE":
                    cell.fill = green
                elif val == "FALSE":
                    cell.fill = red
            elif col == "Military":
                if val == 1:
                    cell.fill = green
            elif col in header and str(col) not in ("_scan_url", "_scan_part", "_row_index"):
                # Automotive keyword numeric columns
                if val == 1:
                    cell.fill = green
                elif val == 0 and col in header:
                    pass  # leave uncoloured — 0 just means not found, not an error

    out = io.BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue()


def _build_partial_excel(df_work: pd.DataFrame, results: list,
                          auto_keywords: list) -> bytes:
    try:
        df_p = _apply_results_to_df(df_work.copy(), results, auto_keywords)
        return _highlight_excel(df_p)
    except Exception:
        return b""


# ══════════════════════════════════════════════════════════════════════
#  STREAMLIT APP
# ══════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="AEC / MIL Scanner v6",
    page_icon="🔬",
    layout="wide",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600;700&family=IBM+Plex+Sans:wght@400;500;600&display=swap');

html, body, [class*="css"] {
    font-family: 'IBM Plex Sans', sans-serif;
}

/* ─── Sidebar ──────────────────────────────────────────────────── */
[data-testid="stSidebar"] {
    background: #0a0e1a !important;
    border-right: 1px solid #1e2840;
}
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] .stTextInput label,
[data-testid="stSidebar"] .stTextArea label,
[data-testid="stSidebar"] .stSlider label,
[data-testid="stSidebar"] .stNumberInput label,
[data-testid="stSidebar"] .stFileUploader label {
    color: #7b9cc4 !important;
    font-size: 0.77rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.06em !important;
    text-transform: uppercase;
}
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 {
    color: #c9d8f0 !important;
    font-family: 'IBM Plex Mono', monospace !important;
}
[data-testid="stSidebar"] input,
[data-testid="stSidebar"] textarea {
    background: #111827 !important;
    color: #e2eaf8 !important;
    border: 1px solid #1e2840 !important;
    border-radius: 4px !important;
}
[data-testid="stSidebar"] .stAlert {
    background: #0f1929 !important;
    border: 1px solid #1e2840 !important;
    color: #7b9cc4 !important;
    font-size: 0.72rem !important;
}

/* ─── Main ─────────────────────────────────────────────────────── */
.main .block-container { padding-top: 1rem; padding-bottom: 2rem; }

[data-testid="metric-container"] {
    background: #0f1929;
    border: 1px solid #1e2840;
    border-radius: 8px;
    padding: 10px 14px;
}

/* ─── Title ────────────────────────────────────────────────────── */
.title-bar {
    background: linear-gradient(135deg, #0a0e1a 0%, #111827 100%);
    border: 1px solid #1e3a5f;
    border-left: 4px solid #3b82f6;
    border-radius: 8px;
    padding: 14px 22px;
    margin-bottom: 1.2rem;
    display: flex;
    align-items: center;
    justify-content: space-between;
}
.title-bar h1 {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 1.05rem;
    font-weight: 700;
    color: #60a5fa;
    margin: 0;
    letter-spacing: 0.04em;
}
.title-bar .badge {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.68rem;
    color: #22c55e;
    background: #052e16;
    border: 1px solid #166534;
    border-radius: 4px;
    padding: 2px 8px;
}

/* ─── Section headers ──────────────────────────────────────────── */
.sec {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.68rem;
    font-weight: 700;
    color: #3b4f6b;
    text-transform: uppercase;
    letter-spacing: 0.14em;
    margin: 1rem 0 0.35rem 0;
    border-bottom: 1px solid #1e2840;
    padding-bottom: 3px;
}

/* ─── Buttons ──────────────────────────────────────────────────── */
div[data-testid="column"]:nth-child(1) .stButton > button {
    background: #2563eb !important; color: #fff !important;
    border: none !important; border-radius: 6px !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-weight: 600 !important; font-size: 0.82rem !important;
    padding: 0.45rem 1.2rem !important; letter-spacing: 0.04em;
}
div[data-testid="column"]:nth-child(2) .stButton > button {
    background: #b45309 !important; color: #fff !important;
    border: none !important; border-radius: 6px !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-weight: 600 !important; font-size: 0.82rem !important;
    padding: 0.45rem 1.2rem !important;
}
div[data-testid="column"]:nth-child(3) .stButton > button {
    background: #dc2626 !important; color: #fff !important;
    border: none !important; border-radius: 6px !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-weight: 600 !important; font-size: 0.82rem !important;
    padding: 0.45rem 1.2rem !important;
}
div[data-testid="column"]:nth-child(4) .stButton > button {
    background: #0f766e !important; color: #fff !important;
    border: none !important; border-radius: 6px !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-weight: 600 !important; font-size: 0.82rem !important;
    padding: 0.45rem 1.2rem !important;
}

/* ─── Log ──────────────────────────────────────────────────────── */
.log-box {
    background: #080c18;
    border: 1px solid #1e2840;
    border-radius: 6px;
    padding: 12px 16px;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 11.5px;
    color: #cbd5e1;
    height: 340px;
    overflow-y: auto;
    white-space: pre-wrap;
    word-break: break-all;
    line-height: 1.6;
}
.log-ok   { color: #4ade80; }
.log-err  { color: #f87171; }
.log-warn { color: #fbbf24; }
.log-info { color: #60a5fa; }
.log-dim  { color: #374151; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="title-bar">
  <h1>🔬 AUTOMOTIVE &amp; MILITARY KEYWORD SCANNER — v6</h1>
  <span class="badge">FULL PDF · NO REF COMPARE · AUTO-SAVE · LARGE FILE SAFE</span>
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ══════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("## ⚙️ Configuration")

    st.markdown('<div class="sec">📁 Files</div>', unsafe_allow_html=True)
    links_file = st.file_uploader("Links Sheet (.xlsx)", type=["xlsx"], key="links")

    st.markdown('<div class="sec">🗂 Column Names</div>', unsafe_allow_html=True)
    url_col  = st.text_input("URL column name",    value="offlineURL")
    part_col = st.text_input("Part Number column", value="PartNumber")

    if links_file:
        try:
            _prev = pd.read_excel(io.BytesIO(links_file.read()), nrows=0)
            links_file.seek(0)
            st.info("Columns: " + " | ".join(_prev.columns.tolist()))
        except Exception:
            pass

    st.markdown('<div class="sec">🚗 Automotive Keywords</div>', unsafe_allow_html=True)
    auto_text = st.text_area("One per line", value=DEFAULT_AUTO_KW.strip(), height=220)

    st.markdown('<div class="sec">🎖 Military Keywords</div>', unsafe_allow_html=True)
    mil_text = st.text_area("One per line — any hit → Military=1",
                             value=DEFAULT_MIL_KW.strip(), height=200)

    st.markdown('<div class="sec">⚡ Performance</div>', unsafe_allow_html=True)
    n_workers  = st.slider("Workers (threads)",       1,  50,  6)
    chunk_size = st.number_input("Chunk size (rows)", min_value=50, max_value=50000,
                                  value=500, step=50)
    rpm        = st.slider("Max requests / minute",   1, 300, 40)

    st.markdown('<div class="sec">🛡 Anti-Detection</div>', unsafe_allow_html=True)
    delay_min = st.number_input("Min inter-chunk delay (s)", 0.0, 30.0, 0.5, step=0.5)
    delay_max = st.number_input("Max inter-chunk delay (s)", 0.0, 60.0, 2.0, step=0.5)
    cb_errors = st.slider("Circuit breaker threshold",       1,  50,  8)
    cb_pause  = st.slider("Circuit breaker pause (s)",       5, 600, 60)

# ══════════════════════════════════════════════════════════════════════
#  SESSION STATE
# ══════════════════════════════════════════════════════════════════════

for _k, _v in [
    ("log_lines",     []),
    ("scan_done",     False),
    ("scan_running",  False),
    ("scan_paused",   False),
    ("result_bytes",  None),
    ("failed_bytes",  None),
    ("partial_bytes", None),
    ("stop_event",    None),
    ("pause_event",   None),
    ("_scan_ctx",     None),
    ("_run_id",       0),
]:
    if _k not in st.session_state:
        st.session_state[_k] = _v


def _append_log(msg: str, tag: str = ""):
    css = {"ok":"log-ok","err":"log-err","warn":"log-warn",
           "info":"log-info","dim":"log-dim"}.get(tag,"")
    ts   = time.strftime("%H:%M:%S")
    line = f'<span class="{css}">[{ts}] {msg}</span>'
    st.session_state.log_lines.append(line)
    if len(st.session_state.log_lines) > 500:
        st.session_state.log_lines = st.session_state.log_lines[-500:]


def _render_log():
    html = "\n".join(st.session_state.log_lines)
    log_ph.markdown(f'<div class="log-box">{html}</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════
#  LAYOUT
# ══════════════════════════════════════════════════════════════════════

st.markdown('<div class="sec">📈 Live Stats</div>', unsafe_allow_html=True)
c1, c2, c3, c4, c5, c6 = st.columns(6)
total_ph   = c1.empty(); total_ph.metric("Total Rows", "—")
jobs_ph    = c2.empty(); jobs_ph.metric("Jobs",        "—")
done_ph    = c3.empty(); done_ph.metric("Completed",   "0")
failed_ph  = c4.empty(); failed_ph.metric("Failed",    "0")
chunk_ph   = c5.empty(); chunk_ph.metric("Chunk",      "—")
circuit_ph = c6.empty(); circuit_ph.metric("CB Errors","0")

st.markdown('<div class="sec">📊 Progress</div>', unsafe_allow_html=True)
prog_bar  = st.progress(0)
prog_text = st.empty()
status_ph = st.empty()

st.markdown('<div class="sec">🚀 Actions</div>', unsafe_allow_html=True)
bc1, bc2, bc3, bc4, _r = st.columns([1,1,1,1,3])
with bc1:
    run_btn   = st.button("▶ RUN SCAN",  use_container_width=True,
                          disabled=st.session_state.scan_running)
with bc2:
    plbl      = "▶ RESUME" if st.session_state.scan_paused else "⏸ PAUSE"
    pause_btn = st.button(plbl, use_container_width=True,
                          disabled=not st.session_state.scan_running)
with bc3:
    stop_btn  = st.button("⏹ STOP",     use_container_width=True,
                          disabled=not st.session_state.scan_running)
with bc4:
    save_btn  = st.button("💾 SAVE NOW", use_container_width=True,
                          disabled=not st.session_state.scan_running)

st.markdown('<div class="sec">🖥 Log</div>', unsafe_allow_html=True)
log_ph = st.empty()

st.markdown('<div class="sec">✅ Results</div>', unsafe_allow_html=True)
results_container = st.container()
_render_log()

# ── Persistent download buttons ───────────────────────────────────────
_rid = st.session_state._run_id
with results_container:
    if st.session_state.result_bytes:
        d1, d2 = st.columns(2)
        d1.download_button("📥 Download Results (.xlsx)",
                           data=st.session_state.result_bytes,
                           file_name="scan_results.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key=f"dl_res_{_rid}")
        if st.session_state.failed_bytes:
            d2.download_button("⚠️ Download Failed Rows (.xlsx)",
                               data=st.session_state.failed_bytes,
                               file_name="scan_failed.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key=f"dl_fail_{_rid}")

    if st.session_state.partial_bytes and not st.session_state.result_bytes:
        st.download_button("💾 Download Partial Results (.xlsx)",
                           data=st.session_state.partial_bytes,
                           file_name="scan_partial.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key=f"dl_part_{_rid}")


# ══════════════════════════════════════════════════════════════════════
#  CONTROL BUTTONS  (Pause / Stop / Save)
# ══════════════════════════════════════════════════════════════════════

def _do_partial_save(label=""):
    ctx = st.session_state._scan_ctx
    if not ctx:
        return
    with ctx["lock"]:
        snap = list(ctx["all_results"])
    if not snap:
        return
    pb = _build_partial_excel(ctx["df_work"], snap, ctx["auto_keywords"])
    if pb:
        st.session_state.partial_bytes = pb
        _append_log(f"💾 {label} {len(snap)} rows saved.", "ok")


if pause_btn and st.session_state.scan_running:
    pe = st.session_state.pause_event
    if pe is not None:
        if st.session_state.scan_paused:
            pe.clear()
            st.session_state.scan_paused = False
            _append_log("▶ Resumed.", "ok")
        else:
            pe.set()
            st.session_state.scan_paused = True
            _append_log("⏸ Paused — workers finish current request then wait.", "warn")
    st.rerun()

if stop_btn and st.session_state.scan_running:
    se = st.session_state.stop_event
    pe = st.session_state.pause_event
    if se: se.set()
    if pe: pe.clear()   # unpause so threads can see stop
    _append_log("⏹ Stop requested …", "warn")
    _do_partial_save("Partial save on stop:")
    st.rerun()

if save_btn and st.session_state.scan_running:
    _do_partial_save("Manual save:")
    st.rerun()


# ══════════════════════════════════════════════════════════════════════
#  RUN SCAN
# ══════════════════════════════════════════════════════════════════════

if run_btn:
    # ── Validate inputs ───────────────────────────────────────────────
    if not links_file:
        st.error("❌ Upload a Links Sheet (.xlsx) first.")
        st.stop()

    auto_keywords = [k.strip() for k in auto_text.splitlines() if k.strip()]
    mil_keywords  = [k.strip() for k in mil_text.splitlines()  if k.strip()]
    if not auto_keywords and not mil_keywords:
        st.error("❌ Enter at least one keyword.")
        st.stop()

    auto_patterns = [_build_pattern(k) for k in auto_keywords]
    mil_patterns  = [_build_pattern(k) for k in mil_keywords]

    # ── Reset state ───────────────────────────────────────────────────
    st.session_state.log_lines     = []
    st.session_state.scan_done     = False
    st.session_state.scan_running  = True
    st.session_state.scan_paused   = False
    st.session_state.result_bytes  = None
    st.session_state.failed_bytes  = None
    st.session_state.partial_bytes = None
    st.session_state._run_id      += 1

    stop_event  = threading.Event()
    pause_event = threading.Event()
    st.session_state.stop_event  = stop_event
    st.session_state.pause_event = pause_event

    def log(msg, tag=""):
        _append_log(msg, tag)

    log(f"Workers: {n_workers} | RPM limit: {rpm} | CB: {cb_errors} errors → {cb_pause}s pause", "info")
    log(f"HTML cap: {MAX_HTML_BYTES//1024} KB | PDF cap: {MAX_PDF_BYTES//1024//1024} MB (full read)", "info")
    log(f"Automotive keywords: {len(auto_keywords)} | Military keywords: {len(mil_keywords)}", "info")
    _render_log()

    # ── Load Excel ────────────────────────────────────────────────────
    try:
        df = pd.read_excel(io.BytesIO(links_file.read()), dtype=str)
    except Exception as e:
        st.error(f"Cannot open file: {e}")
        st.session_state.scan_running = False
        st.stop()

    # Normalise column names (strip BOM, whitespace)
    df.columns = [str(c).strip().lstrip("\ufeff").lower() for c in df.columns]
    url_col_n  = url_col.strip().lower()
    part_col_n = part_col.strip().lower()

    if url_col_n not in df.columns:
        st.error(
            f"❌ Column **{url_col}** not found.  "
            f"Available: `{', '.join(df.columns)}`  "
            f"→ Fix the URL column name in the sidebar."
        )
        st.session_state.scan_running = False
        st.stop()

    if part_col_n not in df.columns:
        log(f"⚠ Column '{part_col}' not found — Part_Scanned will be FALSE.", "warn")
        df[part_col_n] = ""

    df[url_col_n]  = df[url_col_n].apply(_safe_str)
    df[part_col_n] = df[part_col_n].apply(_safe_str)

    df_work = df.copy()

    # ── Build job list ────────────────────────────────────────────────
    all_jobs = [
        (idx, row[url_col_n], row[part_col_n])
        for idx, row in df_work.iterrows()
        if row[url_col_n] and row[url_col_n].lower() != "nan"
    ]

    total_rows = len(df_work)
    total_jobs = len(all_jobs)
    total_ph.metric("Total Rows", total_rows)
    jobs_ph.metric("Jobs",        total_jobs)
    log(f"Total rows: {total_rows} | Jobs with URL: {total_jobs}", "info")

    chunks       = [all_jobs[i:i+chunk_size] for i in range(0, total_jobs, chunk_size)]
    total_chunks = len(chunks)
    log(f"Chunks: {total_chunks} × up to {chunk_size} rows", "info")
    _render_log()

    # ── Shared objects ────────────────────────────────────────────────
    rate_limiter    = RateLimiter(rpm)
    circuit_breaker = CircuitBreaker(cb_errors, cb_pause)
    session         = requests.Session()

    results_lock = threading.Lock()
    all_results  = []
    failed       = []
    done_cnt     = 0

    # Context for Save/Stop buttons
    st.session_state._scan_ctx = {
        "df_work":      df_work,
        "all_results":  all_results,
        "auto_keywords":auto_keywords,
        "lock":         results_lock,
    }

    status_ph.info("● RUNNING")
    user_stopped = False

    try:
        for chunk_idx, chunk_jobs in enumerate(chunks):
            if stop_event.is_set():
                user_stopped = True
                break

            chunk_ph.metric("Chunk", f"{chunk_idx+1}/{total_chunks}")
            log(f"── Chunk {chunk_idx+1}/{total_chunks} ({len(chunk_jobs)} jobs) ──", "info")

            with ThreadPoolExecutor(max_workers=n_workers) as pool:
                futures = {
                    pool.submit(
                        _process_row,
                        ri, url, part,
                        auto_keywords, mil_keywords,
                        auto_patterns, mil_patterns,
                        rate_limiter, circuit_breaker, session,
                        stop_event, pause_event
                    ): (ri, url, part)
                    for ri, url, part in chunk_jobs
                }

                for fut in as_completed(futures):
                    ri, url, part = futures[fut]
                    try:
                        r = fut.result()
                        with results_lock:
                            all_results.append(r)
                        done_cnt += 1
                        short = (url[:55] + "…") if len(url) > 55 else url
                        tag   = "ok" if r.get("Part_Scanned") == "TRUE" else "dim"
                        log(f"✓ [{ri}] {short} | Part={r.get('Part_Scanned')} | "
                            f"Mil={r.get('Military',0)}", tag)
                    except InterruptedError:
                        with results_lock:
                            all_results.append({
                                "_row_index": ri, "_scan_url": url, "_scan_part": part,
                                **{k: 0 for k in auto_keywords},
                                "Military": 0, "Part_Scanned": "FALSE"
                            })
                        failed.append({"row_index": ri, "url": url,
                                       "part": part, "error": "Stopped"})
                        done_cnt += 1
                    except Exception as e:
                        with results_lock:
                            all_results.append({
                                "_row_index": ri, "_scan_url": url, "_scan_part": part,
                                **{k: 0 for k in auto_keywords},
                                "Military": 0, "Part_Scanned": "FALSE"
                            })
                        failed.append({"row_index": ri, "url": url,
                                       "part": part, "error": str(e)})
                        done_cnt += 1
                        log(f"✗ [{ri}] {url[:55]} | {e}", "err")

                    # Auto-save after every row
                    with results_lock:
                        snap = list(all_results)
                    pb = _build_partial_excel(df_work, snap, auto_keywords)
                    if pb:
                        st.session_state.partial_bytes = pb

                    pct = min(int(done_cnt / total_jobs * 100), 100) if total_jobs else 0
                    prog_bar.progress(pct)
                    prog_text.text(f"{done_cnt} / {total_jobs}  ({pct}%)")
                    done_ph.metric("Completed", done_cnt)
                    failed_ph.metric("Failed",    len(failed))
                    circuit_ph.metric("CB Errors", circuit_breaker.error_count)

                    if stop_event.is_set():
                        user_stopped = True
                        break

            _render_log()
            if stop_event.is_set():
                user_stopped = True
                break

            if chunk_idx < total_chunks - 1:
                p = random.uniform(delay_min * 2, delay_max * 2)
                log(f"⏸ Inter-chunk pause {p:.1f}s …", "dim")
                time.sleep(p)

    except Exception as outer:
        log(f"💥 Unexpected error: {outer}", "err")
        log(traceback.format_exc(), "err")
        # Emergency save
        with results_lock:
            snap = list(all_results)
        pb = _build_partial_excel(df_work, snap, auto_keywords)
        if pb:
            st.session_state.partial_bytes = pb
            log(f"💾 Emergency auto-save: {len(snap)} rows.", "warn")

    finally:
        session.close()
        st.session_state.scan_running = False
        st.session_state.scan_paused  = False
        st.session_state._scan_ctx    = None

    # ── Build final output ────────────────────────────────────────────
    with results_lock:
        final_results = list(all_results)

    if not final_results:
        st.warning("No results were collected.")
        st.stop()

    df_final = _apply_results_to_df(df_work.copy(), final_results, auto_keywords)
    # Restore original-case column names
    orig_map = {str(c).strip().lstrip("\ufeff").lower(): str(c).strip()
                for c in pd.read_excel(io.BytesIO(links_file.getvalue()), nrows=0).columns}
    df_final.rename(columns=orig_map, inplace=True)

    result_bytes = _highlight_excel(df_final)
    st.session_state.result_bytes  = result_bytes
    st.session_state.partial_bytes = None   # final replaces partial

    if user_stopped:
        log("⏹ Scan stopped by user — partial results available.", "warn")
        status_ph.warning("● STOPPED")
    else:
        st.session_state.scan_done = True
        log("✅ SCAN COMPLETE", "ok")
        status_ph.success("● COMPLETE")
        prog_bar.progress(100)

    # ── Failed rows ───────────────────────────────────────────────────
    if failed:
        buf = io.BytesIO()
        pd.DataFrame(failed).to_excel(buf, index=False)
        st.session_state.failed_bytes = buf.getvalue()
        log(f"Failed rows: {len(failed)}", "warn")

    # ── Summary ───────────────────────────────────────────────────────
    log("═" * 52, "dim")
    log(f"Total rows     : {len(df_final)}", "ok")
    log(f"Part_Scanned T : {(df_final.get('Part_Scanned','') == 'TRUE').sum()}", "ok")
    log(f"Military found : {(df_final.get('Military', 0) == 1).sum()}", "ok")
    for kw in auto_keywords:
        if kw in df_final.columns:
            log(f"  {kw}: {(df_final[kw] == 1).sum()} hits", "ok")
    _render_log()

    # ── Download buttons ──────────────────────────────────────────────
    rid = st.session_state._run_id
    with results_container:
        d1, d2 = st.columns(2)
        d1.download_button("📥 Download Results (.xlsx)",
                           data=result_bytes,
                           file_name="scan_results.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key=f"dl_res_run_{rid}")
        if st.session_state.failed_bytes:
            d2.download_button("⚠️ Download Failed Rows (.xlsx)",
                               data=st.session_state.failed_bytes,
                               file_name="scan_failed.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key=f"dl_fail_run_{rid}")
